import openpyxl
from openpyxl import Workbook
from typing import Sequence
import string
from pathlib import Path



class XlsxFile:
    COLUMN_NAMES = ('id', 'name', 'email', 'sport', 'location',)

    def __init__(self, filepath: str) -> None:
        self._filepath = filepath
    


    def create(self) -> None:
        """
        Create a XLSX file with filename taken from self._filepath.
        """
        Workbook().save(self._filepath)



    def add_row(self, items: Sequence, row_number: int = 0) -> None:
        """
        Insert data in row with row_number (1, 2, 3,....). if row_number is 0 adds to the end of file.
        """
        # create filepath and file if it does not exist.
        self._prepare_file()

        wb = openpyxl.load_workbook(self._filepath)

        # open active sheet
        sheet = wb.active

        if not sheet:
            print('can not find sheet. add row failed.')
            return

        if row_number == 0:
            row_number = len(self.data()) + 1
        
        letters = self._letters()

        for i, item in enumerate(items):
            address = letters[i] + str(row_number)
            sheet[address] = item

        wb.save(self._filepath)



    def add_rows(self, rows: Sequence[tuple]) -> None:
        """
        Add rows to xlsx file. If file does not exist create new.
        """
        for row in rows:
            self.add_row(items=row, row_number=0)
    


    def data(self) -> tuple:
        """
        Return the content of file by self._filepath.
        """
        # load xlsx file
        wb = openpyxl.load_workbook(self._filepath)

        # open active sheet
        sheet = wb.active
        
        if not sheet:
            print('can not find sheet.')
            return tuple()
        
        return tuple([row for row in sheet.iter_rows(values_only=True)])

    


    def _prepare_file(self) -> None:
        """
        Validate existing of path and file. If doesn't exist create folders and file.
        """
        path = Path(self._filepath)
        
        path.parent.mkdir(parents=True, exist_ok=True)

        if not path.exists():
            self.create()
            self.add_row(self.COLUMN_NAMES)


    
    def _letters(self) -> list[str]:
        """
        Return list of latin uppercase alphabet.
        """
        return [letter for letter in string.ascii_uppercase]

